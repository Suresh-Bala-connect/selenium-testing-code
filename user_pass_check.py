# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# import openpyxl
# import time

# path="C:/Users/Admin/Desktop/testdata.xlsx"
# workbook=openpyxl.load_workbook(path)
# sheet=workbook.active

# service = Service("C:/selenium_drivers/chromedriver.exe")
# driver = webdriver.Chrome(service=service)

# driver.get("https://demo.guru99.com/test/newtours/register.php")
# driver.maximize_window()

# for row in range (2,sheet.max_row + 1):
#     firstname=sheet.cell(row,1).value
#     lastname=sheet.cell(row,2).value
    
#     try:
#         driver.find_element("name","Firstname").clear()
#         driver.find_element("name","Firstname").send_keys(firstname)
        
#         driver.find_element("name","Lastname").clear()
#         driver.find_element("name","Lastname").send_keys(lastname)
        
#         time.sleep(2)
#         sheet.cell(row,4).value="pass"
        
#     except:
#          sheet.cell(row,4).value="fail"
         
#          workbook.save(path)
#          driver.quit
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import openpyxl
import time

# Excel path (make sure this is correct)
path = "C:/Users/Admin/Desktop/testdata.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active

# Setup Selenium
service = Service("C:/selenium_drivers/chromedriver.exe")
driver = webdriver.Chrome(service=service)
driver.maximize_window()
driver.implicitly_wait(5)

for row in range(2, sheet.max_row + 1):
    user = str(sheet.cell(row, 1).value).strip()
    pwd = str(sheet.cell(row, 2).value).strip()

    driver.get("https://www.saucedemo.com/")

    driver.find_element("id", "user-name").send_keys(user)
    driver.find_element("id", "password").send_keys(pwd)
    driver.find_element("id", "login-button").click()

    if "inventory" in driver.current_url:
        sheet.cell(row, 3).value = "PASS"
    else:
        sheet.cell(row, 3).value = "FAIL"

# ✅ Save before quitting
wb.save(path)
driver.quit()

print("✅ Login test completed — PASS/FAIL written to Excel.")

