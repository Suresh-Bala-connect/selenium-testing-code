from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time

# Config - easier to modify
EXCEL_PATH = "C:/Users/Admin/Desktop/testdata.xlsx"
DRIVER_PATH = "C:/selenium_drivers/chromedriver.exe"
LOGIN_URL = "file:///D:/AnjanaInfoTech/Python/test.html"


# Load Excel
wb = openpyxl.load_workbook("C:/Users/Admin/Desktop/userpass.xlsx")
sheet = wb.active

# Add result header if not present
if sheet.max_column < 3:
    sheet.cell(1, 3, "Result")

# Setup WebDriver
service = Service("c:/selenium_drivers/chromedriver.exe")
driver = webdriver.Chrome(service=service)
driver.maximize_window()

try:
    for row in range(2, sheet.max_row + 1):
        user = str(sheet.cell(row, 1).value).strip()
        pwd = str(sheet.cell(row, 2).value).strip()

        driver.get(LOGIN_URL)
        
        
        # Enter credentials
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "username"))
        ).send_keys(user)
        
        driver.find_element(By.ID, "password").send_keys(pwd)
        driver.find_element(By.XPATH, '//button[text()="Login"]').click()

        # Verify login success (register form should appear)
        try:
            WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.ID, "registerForm"))
            )
            sheet.cell(row, 3).value = "PASS"
            print(f"✅ {user}: PASS")
            
            # Optional: Take screenshot for evidence
            # driver.save_screenshot(f"success_{user}.png")
        except:
            sheet.cell(row, 3).value = "FAIL"
            print(f"❌ {user}: FAIL")
            
            # Optional: Capture error message if available
            # error = driver.find_element(By.CSS_SELECTOR, ".error-message").text
            # print(f"Error: {error}")
finally:
    # Ensure resources are cleaned up
    wb.save("userpass_results.xlsx")
    driver.quit()
    print("✅ Login test completed — PASS/FAIL written to Excel.")