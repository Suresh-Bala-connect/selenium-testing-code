import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import time

# Load Excel file
workbook = openpyxl.load_workbook("C:/Users/Admin/Desktop/book1.xlsx")
sheet = workbook.active  # First sheet

# Read values and handle None
firstname = (sheet.cell(row=5, column=4).value or "").strip()
lastname = (sheet.cell(row=5, column=5).value or "").strip()

print(firstname, lastname)  # Should print: Suresh Bala


# Start Chrome
service = Service("C:\selenium_drivers\chromedriver.exe")
driver = webdriver.Chrome(service=service)

driver.get("https://www.techlistic.com/p/selenium-practice-form.html")
driver.maximize_window()
time.sleep(2)

# Fill First Name
driver.find_element("name", "firstname").send_keys(firstname)
time.sleep(1)

# Fill Last Name
driver.find_element("name", "lastname").send_keys(lastname)
time.sleep(1)


# ✅ Add more fields if needed...
print("Form filled successfully!")

time.sleep(3)
driver.quit()


# # File path
# path = "C:/Users/Admin/Desktop/book1.xlsx"

# # Load workbook
# workbook = openpyxl.load_workbook(path)
# sheet = workbook.active

# # Example: Directly write results
# sheet["D2"] = "PASS"   # Row 2, Column D
# sheet["D3"] = "FAIL"   # Row 3, Column D

# # Save changes
# workbook.save(path)

# print("✅ Results updated successfully!")
