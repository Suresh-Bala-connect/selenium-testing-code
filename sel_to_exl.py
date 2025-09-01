from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import openpyxl
import time

# Load Excel file
path = "C:/Users/Admin/Desktop/book1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

# Setup Chrome
service = Service("C:/selenium_drivers/chromedriver.exe")
driver = webdriver.Chrome(service=service)

driver.get("https://www.techlistic.com/p/selenium-practice-form.html")
driver.maximize_window()

# Loop through rows and write results
for row in range(8, sheet.max_row + 1):
    firstname = sheet.cell(row, 1).value
    lastname = sheet.cell(row, 2).value

    try:
        driver.find_element("name", "firstname").clear()
        driver.find_element("name", "firstname").send_keys(firstname)

        driver.find_element("name", "lastname").clear()
        driver.find_element("name", "lastname").send_keys(lastname)

        time.sleep(2)

        # If no errors, mark as PASS
        sheet.cell(row, 4).value = "PASS"
    except:
        # If error occurs, mark as FAIL
        sheet.cell(row, 4).value = "FAIL"

# Save the file
workbook.save(path)

driver.quit()
